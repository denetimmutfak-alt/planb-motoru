"""
PlanB Motoru - Report Generator
Excel/PDF rapor çıktıları
"""
import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime, timedelta
import os
from src.utils.logger import log_info, log_error, log_debug

class ReportGenerator:
    """Rapor oluşturucu"""
    
    def __init__(self):
        self.export_directory = "data/exports"
        self._ensure_export_directory()
        
        # Excel için gerekli kütüphaneler
        self.excel_available = self._check_excel_dependencies()
        
        # PDF için gerekli kütüphaneler
        self.pdf_available = self._check_pdf_dependencies()
    
    def _ensure_export_directory(self):
        """Export dizinini oluştur"""
        os.makedirs(self.export_directory, exist_ok=True)
        os.makedirs(f"{self.export_directory}/excel", exist_ok=True)
        os.makedirs(f"{self.export_directory}/pdf", exist_ok=True)
    
    def _check_excel_dependencies(self) -> bool:
        """Excel bağımlılıklarını kontrol et"""
        try:
            import openpyxl
            return True
        except ImportError:
            log_error("openpyxl kütüphanesi bulunamadı. Excel export için gerekli.")
            return False
    
    def _check_pdf_dependencies(self) -> bool:
        """PDF bağımlılıklarını kontrol et"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            return True
        except ImportError:
            log_error("reportlab kütüphanesi bulunamadı. PDF export için gerekli.")
            return False
    
    def generate_portfolio_report(self, portfolio_name: str, format: str = 'excel') -> Optional[str]:
        """Portföy raporu oluştur"""
        try:
            from src.portfolio.portfolio_manager import portfolio_manager
            
            portfolio = portfolio_manager.get_portfolio(portfolio_name)
            if not portfolio:
                log_error(f"Portföy bulunamadı: {portfolio_name}")
                return None
            
            # Portföy verilerini hazırla
            portfolio_data = self._prepare_portfolio_data(portfolio)
            
            if format.lower() == 'excel':
                return self._generate_excel_portfolio_report(portfolio_name, portfolio_data)
            elif format.lower() == 'pdf':
                return self._generate_pdf_portfolio_report(portfolio_name, portfolio_data)
            else:
                log_error(f"Desteklenmeyen format: {format}")
                return None
                
        except Exception as e:
            log_error(f"Portföy raporu oluşturma hatası: {e}")
            return None
    
    def generate_analysis_report(self, symbol: str, analysis_data: Dict[str, Any], format: str = 'excel') -> Optional[str]:
        """Analiz raporu oluştur"""
        try:
            # Analiz verilerini hazırla
            report_data = self._prepare_analysis_data(symbol, analysis_data)
            
            if format.lower() == 'excel':
                return self._generate_excel_analysis_report(symbol, report_data)
            elif format.lower() == 'pdf':
                return self._generate_pdf_analysis_report(symbol, report_data)
            else:
                log_error(f"Desteklenmeyen format: {format}")
                return None
                
        except Exception as e:
            log_error(f"Analiz raporu oluşturma hatası: {e}")
            return None
    
    def generate_market_report(self, market: str, symbols: List[str], format: str = 'excel') -> Optional[str]:
        """Piyasa raporu oluştur"""
        try:
            # Piyasa verilerini hazırla
            market_data = self._prepare_market_data(market, symbols)
            
            if format.lower() == 'excel':
                return self._generate_excel_market_report(market, market_data)
            elif format.lower() == 'pdf':
                return self._generate_pdf_market_report(market, market_data)
            else:
                log_error(f"Desteklenmeyen format: {format}")
                return None
                
        except Exception as e:
            log_error(f"Piyasa raporu oluşturma hatası: {e}")
            return None
    
    def _prepare_portfolio_data(self, portfolio) -> Dict[str, Any]:
        """Portföy verilerini hazırla"""
        try:
            # Pozisyon verileri
            positions_data = []
            total_value = portfolio.cash
            
            for symbol, position in portfolio.positions.items():
                position_value = position['quantity'] * position['current_price']
                total_value += position_value
                
                pnl = (position['current_price'] - position['avg_price']) * position['quantity']
                pnl_pct = ((position['current_price'] - position['avg_price']) / position['avg_price'] * 100) if position['avg_price'] > 0 else 0
                
                positions_data.append({
                    'Symbol': symbol,
                    'Quantity': position['quantity'],
                    'Avg Price': position['avg_price'],
                    'Current Price': position['current_price'],
                    'Total Value': position_value,
                    'P&L': pnl,
                    'P&L %': pnl_pct,
                    'Market': position.get('market', 'Unknown')
                })
            
            # İşlem geçmişi
            transactions_data = []
            for transaction in portfolio.transactions:
                transactions_data.append({
                    'Date': transaction['date'],
                    'Type': transaction['type'],
                    'Symbol': transaction['symbol'],
                    'Quantity': transaction['quantity'],
                    'Price': transaction['price'],
                    'Total': transaction['quantity'] * transaction['price'],
                    'Commission': transaction.get('commission', 0)
                })
            
            # Portföy özeti
            portfolio_summary = {
                'Portfolio Name': portfolio.name,
                'Total Value': total_value,
                'Cash': portfolio.cash,
                'Invested Value': total_value - portfolio.cash,
                'Total P&L': sum(pos['P&L'] for pos in positions_data),
                'Total P&L %': ((total_value - portfolio.initial_capital) / portfolio.initial_capital * 100) if portfolio.initial_capital > 0 else 0,
                'Position Count': len(portfolio.positions),
                'Transaction Count': len(portfolio.transactions),
                'Created Date': portfolio.created_date,
                'Last Updated': datetime.now().isoformat()
            }
            
            return {
                'summary': portfolio_summary,
                'positions': positions_data,
                'transactions': transactions_data
            }
            
        except Exception as e:
            log_error(f"Portföy veri hazırlama hatası: {e}")
            return {}
    
    def _prepare_analysis_data(self, symbol: str, analysis_data: Dict[str, Any]) -> Dict[str, Any]:
        """Analiz verilerini hazırla"""
        try:
            # Temel analiz verileri
            basic_analysis = {
                'Symbol': symbol,
                'Total Score': analysis_data.get('total_score', 0),
                'Signal': analysis_data.get('signal', 'HOLD'),
                'Confidence': analysis_data.get('confidence', 0),
                'Analysis Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            # Detaylı analiz skorları
            detailed_scores = []
            detailed_analysis = analysis_data.get('detailed_analysis', {})
            
            score_components = [
                ('Technical Analysis', 'technical_score'),
                ('Financial Analysis', 'financial_score'),
                ('Sentiment Analysis', 'sentiment_score'),
                ('Astrology Analysis', 'astrology_score'),
                ('Shemitah Cycle', 'shemitah_score'),
                ('Gann Technique', 'gann_score'),
                ('Spiral21 Cycle', 'spiral21_score'),
                ('Solar Cycle', 'solar_score'),
                ('Moon Phases', 'moon_score'),
                ('Gann-Astro Hybrid', 'gann_astro_score'),
                ('Statistical Validation', 'statistical_score'),
                ('ML Prediction', 'ml_score')
            ]
            
            for component_name, score_key in score_components:
                score = detailed_analysis.get(score_key, 0)
                detailed_scores.append({
                    'Component': component_name,
                    'Score': score,
                    'Weight': analysis_data.get('weights', {}).get(score_key, 0),
                    'Weighted Score': score * analysis_data.get('weights', {}).get(score_key, 0)
                })
            
            # Risk faktörleri
            risk_factors = []
            if 'risk_factors' in detailed_analysis:
                for risk in detailed_analysis['risk_factors']:
                    risk_factors.append({
                        'Risk Type': risk.get('type', 'Unknown'),
                        'Severity': risk.get('severity', 'Medium'),
                        'Description': risk.get('description', ''),
                        'Impact': risk.get('impact', 0)
                    })
            
            # Öneriler
            recommendations = []
            if 'recommendations' in detailed_analysis:
                for rec in detailed_analysis['recommendations']:
                    recommendations.append({
                        'Type': rec.get('type', 'General'),
                        'Priority': rec.get('priority', 'Medium'),
                        'Description': rec.get('description', ''),
                        'Action': rec.get('action', '')
                    })
            
            return {
                'basic_analysis': basic_analysis,
                'detailed_scores': detailed_scores,
                'risk_factors': risk_factors,
                'recommendations': recommendations,
                'raw_analysis': analysis_data
            }
            
        except Exception as e:
            log_error(f"Analiz veri hazırlama hatası: {e}")
            return {}
    
    def _prepare_market_data(self, market: str, symbols: List[str]) -> Dict[str, Any]:
        """Piyasa verilerini hazırla"""
        try:
            # Simüle edilmiş piyasa verileri
            market_data = []
            
            for symbol in symbols:
                # Simüle edilmiş veriler
                price = np.random.uniform(10, 1000)
                change = np.random.uniform(-10, 10)
                volume = np.random.uniform(1000000, 10000000)
                
                market_data.append({
                    'Symbol': symbol,
                    'Price': price,
                    'Change': change,
                    'Change %': change,
                    'Volume': volume,
                    'Market Cap': price * volume * 0.1,  # Simüle edilmiş
                    'PE Ratio': np.random.uniform(5, 50),
                    '52W High': price * 1.2,
                    '52W Low': price * 0.8
                })
            
            # Piyasa özeti
            market_summary = {
                'Market': market,
                'Total Symbols': len(symbols),
                'Average Price': np.mean([data['Price'] for data in market_data]),
                'Average Change': np.mean([data['Change'] for data in market_data]),
                'Total Volume': sum([data['Volume'] for data in market_data]),
                'Report Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            return {
                'summary': market_summary,
                'market_data': market_data
            }
            
        except Exception as e:
            log_error(f"Piyasa veri hazırlama hatası: {e}")
            return {}
    
    def _generate_excel_portfolio_report(self, portfolio_name: str, portfolio_data: Dict[str, Any]) -> Optional[str]:
        """Excel portföy raporu oluştur"""
        try:
            if not self.excel_available:
                log_error("Excel export için gerekli kütüphaneler yüklü değil")
                return None
            
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Excel dosyası oluştur
            filename = f"{self.export_directory}/excel/portfolio_{portfolio_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb = openpyxl.Workbook()
            
            # Portföy özeti sayfası
            ws_summary = wb.active
            ws_summary.title = "Portfolio Summary"
            
            # Başlık
            ws_summary['A1'] = f"{portfolio_name} Portfolio Report"
            ws_summary['A1'].font = Font(size=16, bold=True)
            ws_summary['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            ws_summary['A1'].font = Font(color="FFFFFF", size=16, bold=True)
            
            # Özet verileri
            summary_data = portfolio_data['summary']
            row = 3
            for key, value in summary_data.items():
                ws_summary[f'A{row}'] = key
                ws_summary[f'B{row}'] = value
                row += 1
            
            # Pozisyonlar sayfası
            ws_positions = wb.create_sheet("Positions")
            positions_df = pd.DataFrame(portfolio_data['positions'])
            
            if not positions_df.empty:
                for r in dataframe_to_rows(positions_df, index=False, header=True):
                    ws_positions.append(r)
                
                # Stil uygula
                self._apply_excel_styles(ws_positions, len(positions_df.columns), len(positions_df) + 1)
            
            # İşlemler sayfası
            ws_transactions = wb.create_sheet("Transactions")
            transactions_df = pd.DataFrame(portfolio_data['transactions'])
            
            if not transactions_df.empty:
                for r in dataframe_to_rows(transactions_df, index=False, header=True):
                    ws_transactions.append(r)
                
                # Stil uygula
                self._apply_excel_styles(ws_transactions, len(transactions_df.columns), len(transactions_df) + 1)
            
            # Dosyayı kaydet
            wb.save(filename)
            log_info(f"Excel portföy raporu oluşturuldu: {filename}")
            return filename
            
        except Exception as e:
            log_error(f"Excel portföy raporu oluşturma hatası: {e}")
            return None
    
    def _generate_pdf_portfolio_report(self, portfolio_name: str, portfolio_data: Dict[str, Any]) -> Optional[str]:
        """PDF portföy raporu oluştur"""
        try:
            if not self.pdf_available:
                log_error("PDF export için gerekli kütüphaneler yüklü değil")
                return None
            
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            
            # PDF dosyası oluştur
            filename = f"{self.export_directory}/pdf/portfolio_{portfolio_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            doc = SimpleDocTemplate(filename, pagesize=A4)
            
            # Stil tanımları
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # Center
            )
            
            # İçerik oluştur
            story = []
            
            # Başlık
            story.append(Paragraph(f"{portfolio_name} Portfolio Report", title_style))
            story.append(Spacer(1, 20))
            
            # Portföy özeti
            story.append(Paragraph("Portfolio Summary", styles['Heading2']))
            summary_data = portfolio_data['summary']
            
            summary_table_data = [['Metric', 'Value']]
            for key, value in summary_data.items():
                summary_table_data.append([key, str(value)])
            
            summary_table = Table(summary_table_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 20))
            
            # Pozisyonlar
            if portfolio_data['positions']:
                story.append(Paragraph("Positions", styles['Heading2']))
                
                positions_data = portfolio_data['positions']
                positions_table_data = [['Symbol', 'Quantity', 'Avg Price', 'Current Price', 'Total Value', 'P&L', 'P&L %']]
                
                for pos in positions_data:
                    positions_table_data.append([
                        pos['Symbol'],
                        f"{pos['Quantity']:.2f}",
                        f"{pos['Avg Price']:.2f}",
                        f"{pos['Current Price']:.2f}",
                        f"{pos['Total Value']:.2f}",
                        f"{pos['P&L']:.2f}",
                        f"{pos['P&L %']:.2f}%"
                    ])
                
                positions_table = Table(positions_table_data)
                positions_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8)
                ]))
                
                story.append(positions_table)
                story.append(PageBreak())
            
            # İşlemler
            if portfolio_data['transactions']:
                story.append(Paragraph("Transactions", styles['Heading2']))
                
                transactions_data = portfolio_data['transactions']
                transactions_table_data = [['Date', 'Type', 'Symbol', 'Quantity', 'Price', 'Total', 'Commission']]
                
                for trans in transactions_data:
                    transactions_table_data.append([
                        trans['Date'],
                        trans['Type'],
                        trans['Symbol'],
                        f"{trans['Quantity']:.2f}",
                        f"{trans['Price']:.2f}",
                        f"{trans['Total']:.2f}",
                        f"{trans['Commission']:.2f}"
                    ])
                
                transactions_table = Table(transactions_table_data)
                transactions_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8)
                ]))
                
                story.append(transactions_table)
            
            # PDF oluştur
            doc.build(story)
            log_info(f"PDF portföy raporu oluşturuldu: {filename}")
            return filename
            
        except Exception as e:
            log_error(f"PDF portföy raporu oluşturma hatası: {e}")
            return None
    
    def _generate_excel_analysis_report(self, symbol: str, report_data: Dict[str, Any]) -> Optional[str]:
        """Excel analiz raporu oluştur"""
        try:
            if not self.excel_available:
                return None
            
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            filename = f"{self.export_directory}/excel/analysis_{symbol}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb = openpyxl.Workbook()
            
            # Temel analiz sayfası
            ws_basic = wb.active
            ws_basic.title = "Basic Analysis"
            
            # Başlık
            ws_basic['A1'] = f"{symbol} Analysis Report"
            ws_basic['A1'].font = Font(size=16, bold=True)
            
            # Temel analiz verileri
            basic_data = report_data['basic_analysis']
            row = 3
            for key, value in basic_data.items():
                ws_basic[f'A{row}'] = key
                ws_basic[f'B{row}'] = value
                row += 1
            
            # Detaylı skorlar sayfası
            ws_scores = wb.create_sheet("Detailed Scores")
            scores_df = pd.DataFrame(report_data['detailed_scores'])
            
            if not scores_df.empty:
                for r in dataframe_to_rows(scores_df, index=False, header=True):
                    ws_scores.append(r)
                
                self._apply_excel_styles(ws_scores, len(scores_df.columns), len(scores_df) + 1)
            
            # Risk faktörleri sayfası
            if report_data['risk_factors']:
                ws_risks = wb.create_sheet("Risk Factors")
                risks_df = pd.DataFrame(report_data['risk_factors'])
                
                for r in dataframe_to_rows(risks_df, index=False, header=True):
                    ws_risks.append(r)
                
                self._apply_excel_styles(ws_risks, len(risks_df.columns), len(risks_df) + 1)
            
            # Öneriler sayfası
            if report_data['recommendations']:
                ws_recommendations = wb.create_sheet("Recommendations")
                recommendations_df = pd.DataFrame(report_data['recommendations'])
                
                for r in dataframe_to_rows(recommendations_df, index=False, header=True):
                    ws_recommendations.append(r)
                
                self._apply_excel_styles(ws_recommendations, len(recommendations_df.columns), len(recommendations_df) + 1)
            
            wb.save(filename)
            log_info(f"Excel analiz raporu oluşturuldu: {filename}")
            return filename
            
        except Exception as e:
            log_error(f"Excel analiz raporu oluşturma hatası: {e}")
            return None
    
    def _generate_pdf_analysis_report(self, symbol: str, report_data: Dict[str, Any]) -> Optional[str]:
        """PDF analiz raporu oluştur"""
        try:
            if not self.pdf_available:
                return None
            
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            
            filename = f"{self.export_directory}/pdf/analysis_{symbol}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            doc = SimpleDocTemplate(filename, pagesize=A4)
            
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1
            )
            
            story = []
            
            # Başlık
            story.append(Paragraph(f"{symbol} Analysis Report", title_style))
            story.append(Spacer(1, 20))
            
            # Temel analiz
            story.append(Paragraph("Basic Analysis", styles['Heading2']))
            basic_data = report_data['basic_analysis']
            
            basic_table_data = [['Metric', 'Value']]
            for key, value in basic_data.items():
                basic_table_data.append([key, str(value)])
            
            basic_table = Table(basic_table_data)
            basic_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(basic_table)
            story.append(Spacer(1, 20))
            
            # Detaylı skorlar
            story.append(Paragraph("Detailed Scores", styles['Heading2']))
            scores_data = report_data['detailed_scores']
            
            scores_table_data = [['Component', 'Score', 'Weight', 'Weighted Score']]
            for score in scores_data:
                scores_table_data.append([
                    score['Component'],
                    f"{score['Score']:.2f}",
                    f"{score['Weight']:.2f}",
                    f"{score['Weighted Score']:.2f}"
                ])
            
            scores_table = Table(scores_table_data)
            scores_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8)
            ]))
            
            story.append(scores_table)
            
            doc.build(story)
            log_info(f"PDF analiz raporu oluşturuldu: {filename}")
            return filename
            
        except Exception as e:
            log_error(f"PDF analiz raporu oluşturma hatası: {e}")
            return None
    
    def _generate_excel_market_report(self, market: str, market_data: Dict[str, Any]) -> Optional[str]:
        """Excel piyasa raporu oluştur"""
        try:
            if not self.excel_available:
                return None
            
            import openpyxl
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            filename = f"{self.export_directory}/excel/market_{market}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb = openpyxl.Workbook()
            
            # Piyasa özeti sayfası
            ws_summary = wb.active
            ws_summary.title = "Market Summary"
            
            # Başlık
            ws_summary['A1'] = f"{market} Market Report"
            ws_summary['A1'].font = Font(size=16, bold=True)
            
            # Piyasa özeti
            summary_data = market_data['summary']
            row = 3
            for key, value in summary_data.items():
                ws_summary[f'A{row}'] = key
                ws_summary[f'B{row}'] = value
                row += 1
            
            # Piyasa verileri sayfası
            ws_market = wb.create_sheet("Market Data")
            market_df = pd.DataFrame(market_data['market_data'])
            
            if not market_df.empty:
                for r in dataframe_to_rows(market_df, index=False, header=True):
                    ws_market.append(r)
                
                self._apply_excel_styles(ws_market, len(market_df.columns), len(market_df) + 1)
            
            wb.save(filename)
            log_info(f"Excel piyasa raporu oluşturuldu: {filename}")
            return filename
            
        except Exception as e:
            log_error(f"Excel piyasa raporu oluşturma hatası: {e}")
            return None
    
    def _generate_pdf_market_report(self, market: str, market_data: Dict[str, Any]) -> Optional[str]:
        """PDF piyasa raporu oluştur"""
        try:
            if not self.pdf_available:
                return None
            
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            
            filename = f"{self.export_directory}/pdf/market_{market}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            doc = SimpleDocTemplate(filename, pagesize=A4)
            
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1
            )
            
            story = []
            
            # Başlık
            story.append(Paragraph(f"{market} Market Report", title_style))
            story.append(Spacer(1, 20))
            
            # Piyasa özeti
            story.append(Paragraph("Market Summary", styles['Heading2']))
            summary_data = market_data['summary']
            
            summary_table_data = [['Metric', 'Value']]
            for key, value in summary_data.items():
                summary_table_data.append([key, str(value)])
            
            summary_table = Table(summary_table_data)
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 20))
            
            # Piyasa verileri
            story.append(Paragraph("Market Data", styles['Heading2']))
            market_data_list = market_data['market_data']
            
            market_table_data = [['Symbol', 'Price', 'Change', 'Change %', 'Volume', 'Market Cap', 'PE Ratio']]
            for data in market_data_list:
                market_table_data.append([
                    data['Symbol'],
                    f"{data['Price']:.2f}",
                    f"{data['Change']:.2f}",
                    f"{data['Change %']:.2f}%",
                    f"{data['Volume']:,.0f}",
                    f"{data['Market Cap']:,.0f}",
                    f"{data['PE Ratio']:.2f}"
                ])
            
            market_table = Table(market_table_data)
            market_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 8)
            ]))
            
            story.append(market_table)
            
            doc.build(story)
            log_info(f"PDF piyasa raporu oluşturuldu: {filename}")
            return filename
            
        except Exception as e:
            log_error(f"PDF piyasa raporu oluşturma hatası: {e}")
            return None
    
    def _apply_excel_styles(self, worksheet, num_cols: int, num_rows: int):
        """Excel stil uygula"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Başlık satırı stili
            for col in range(1, num_cols + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Kenarlık
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(1, num_rows + 1):
                for col in range(1, num_cols + 1):
                    worksheet.cell(row=row, column=col).border = thin_border
            
            # Sütun genişlikleri
            for col in range(1, num_cols + 1):
                worksheet.column_dimensions[worksheet.cell(row=1, column=col).column_letter].width = 15
            
        except Exception as e:
            log_error(f"Excel stil uygulama hatası: {e}")
    
    def get_export_status(self) -> Dict[str, Any]:
        """Export durumunu getir"""
        try:
            return {
                'excel_available': self.excel_available,
                'pdf_available': self.pdf_available,
                'export_directory': self.export_directory,
                'excel_directory': f"{self.export_directory}/excel",
                'pdf_directory': f"{self.export_directory}/pdf",
                'excel_files': len([f for f in os.listdir(f"{self.export_directory}/excel") if f.endswith('.xlsx')]) if os.path.exists(f"{self.export_directory}/excel") else 0,
                'pdf_files': len([f for f in os.listdir(f"{self.export_directory}/pdf") if f.endswith('.pdf')]) if os.path.exists(f"{self.export_directory}/pdf") else 0
            }
        except Exception as e:
            log_error(f"Export durumu alma hatası: {e}")
            return {}

# Global report generator instance
report_generator = ReportGenerator()

